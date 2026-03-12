"""
ROI Report Generator — Fills the ROI/Drive Time Excel template with scraped data.

Copies the template, writes raw metric cells (all ratios auto-compute via formulas).
"""

from __future__ import annotations

import shutil
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

import openpyxl

from roi_scraper import TradeMetrics, DriveTimeMetrics


TEMPLATE_PATH = Path(__file__).parent / "roi_template.xlsx"


# ─── Cell Maps ───────────────────────────────────────────────────────────────

# Basic sheet: trade → (pre_col, post_col) for each metric row
# HVAC: cols G/H, rows 5-12
# Plumbing: cols M/N, rows 5-10 (no Tech Lead Sales or Leads Set)
# Electrical: cols S/T, rows 5-10
# Drains: cols Y/Z, rows 5-12

BASIC_HVAC = {
    "pre_col": "G", "post_col": "H",
    "total_sales": 5,
    "total_tech_lead_sales": 6,
    "completed_revenue": 7,
    "completed_jobs": 8,
    "opportunities": 9,
    "sales_opportunities": 10,
    "leads_set": 11,
    "converted_jobs": 12,
    "date_row": 4,
}

BASIC_PLUMBING = {
    "pre_col": "M", "post_col": "N",
    "total_sales": 5,
    "completed_revenue": 6,
    "completed_jobs": 7,
    "opportunities": 8,
    "sales_opportunities": 9,
    "converted_jobs": 10,
    "date_row": 4,
}

BASIC_ELECTRICAL = {
    "pre_col": "S", "post_col": "T",
    "total_sales": 5,
    "completed_revenue": 6,
    "completed_jobs": 7,
    "opportunities": 8,
    "sales_opportunities": 9,
    "converted_jobs": 10,
    "date_row": 4,
}

BASIC_DRAINS = {
    "pre_col": "Y", "post_col": "Z",
    "total_sales": 5,
    "total_tech_lead_sales": 6,
    "completed_revenue": 7,
    "completed_jobs": 8,
    "opportunities": 9,
    "sales_opportunities": 10,
    "leads_set": 11,
    "converted_jobs": 12,
    "date_row": 4,
}

BASIC_TRADE_MAP = {
    "HVAC": BASIC_HVAC,
    "Plumbing": BASIC_PLUMBING,
    "Electrical": BASIC_ELECTRICAL,
    "Drains": BASIC_DRAINS,
}

# Natural delta: HVAC uses rows 24-31, cols G/H
# Same column layout as PB delta but offset by 19 rows for HVAC
NATURAL_HVAC = {
    "pre_col": "G", "post_col": "H",
    "total_sales": 24,
    "total_tech_lead_sales": 25,
    "completed_revenue": 26,
    "completed_jobs": 27,
    "opportunities": 28,
    "sales_opportunities": 29,
    "leads_set": 30,
    "converted_jobs": 31,
    "date_row": 23,
}

# High Level sheet: cols C/D, rows 5-12 (PB), rows 24-31 (Natural)
HIGH_LEVEL_PB = {
    "pre_col": "C", "post_col": "D",
    "total_sales": 5,
    "total_tech_lead_sales": 6,
    "completed_revenue": 7,
    "completed_jobs": 8,
    "opportunities": 9,
    "sales_opportunities": 10,
    "leads_set": 11,
    "converted_jobs": 12,
    "date_row": 4,
}

HIGH_LEVEL_NATURAL = {
    "pre_col": "C", "post_col": "D",
    "total_sales": 24,
    "total_tech_lead_sales": 25,
    "completed_revenue": 26,
    "completed_jobs": 27,
    "opportunities": 28,
    "sales_opportunities": 29,
    "leads_set": 30,
    "converted_jobs": 31,
    "date_row": 23,
}

# Drive Time sheet: cols C/D, rows 4-8
DRIVE_TIME_MAP = {
    "pre_col": "C", "post_col": "D",
    "total_drives": 4,
    "jobs": 5,
    "drive_time_hours": 6,
    "idle_time_hours": 7,
    "working_time_hours": 8,
    "date_row": 3,
}


# ─── Writer ──────────────────────────────────────────────────────────────────

def _write_metrics(ws, cell_map: dict, pre: TradeMetrics, post: TradeMetrics):
    """Write TradeMetrics pre/post values into worksheet cells."""
    pre_col = cell_map["pre_col"]
    post_col = cell_map["post_col"]

    metric_fields = [
        "total_sales", "total_tech_lead_sales", "completed_revenue",
        "completed_jobs", "opportunities", "sales_opportunities",
        "leads_set", "converted_jobs",
    ]

    pre_dict = asdict(pre)
    post_dict = asdict(post)

    for field_name in metric_fields:
        if field_name not in cell_map:
            continue  # This trade doesn't have this metric (e.g., Plumbing has no leads_set)
        row = cell_map[field_name]
        ws[f"{pre_col}{row}"] = pre_dict[field_name]
        ws[f"{post_col}{row}"] = post_dict[field_name]


def _write_drive_time(ws, pre: DriveTimeMetrics, post: DriveTimeMetrics):
    """Write DriveTimeMetrics pre/post values."""
    m = DRIVE_TIME_MAP
    pre_col = m["pre_col"]
    post_col = m["post_col"]

    for field_name in ["total_drives", "jobs", "drive_time_hours", "idle_time_hours", "working_time_hours"]:
        row = m[field_name]
        ws[f"{pre_col}{row}"] = getattr(pre, field_name)
        ws[f"{post_col}{row}"] = getattr(post, field_name)


def _write_date_labels(ws, cell_map: dict, pre_label: str, post_label: str):
    """Write date range labels."""
    row = cell_map["date_row"]
    ws[f"{cell_map['pre_col']}{row}"] = pre_label
    ws[f"{cell_map['post_col']}{row}"] = post_label


def generate_roi_report(
    data: dict,
    org_name: str,
    pre_start: str,
    pre_end: str,
    post_start: str,
    post_end: str,
    natural_y1_label: str = "",
    natural_y2_label: str = "",
    output_dir: str = ".",
) -> str:
    """
    Fill the ROI template with extracted data.

    Args:
        data: Output from pull_roi_data()
        org_name: Customer name for filename
        pre_start/pre_end: Pre-PB date range labels
        post_start/post_end: Post-PB date range labels
        natural_y1_label/y2_label: Natural delta date labels
        output_dir: Where to save the output

    Returns:
        Path to the generated Excel file.
    """
    # Copy template
    timestamp = datetime.now().strftime("%Y-%m-%d")
    safe_name = org_name.replace("/", "-").replace(" ", "_")
    output_path = Path(output_dir) / f"{safe_name}_ROI_Report_{timestamp}.xlsx"
    shutil.copy2(TEMPLATE_PATH, output_path)

    wb = openpyxl.load_workbook(output_path)

    pre_label = f"{pre_start} - {pre_end}"
    post_label = f"{post_start} - {post_end}"

    # ── Basic Sheet ──
    if "Basic" in wb.sheetnames:
        ws = wb["Basic"]

        for trade, cell_map in BASIC_TRADE_MAP.items():
            if trade in data.get("basic", {}):
                trade_data = data["basic"][trade]
                _write_metrics(ws, cell_map, trade_data["pre"], trade_data["post"])
                _write_date_labels(ws, cell_map, pre_label, post_label)

        # Natural delta for HVAC (rows 24-31)
        if "HVAC" in data.get("natural", {}):
            nat = data["natural"]["HVAC"]
            _write_metrics(ws, NATURAL_HVAC, nat["y1"], nat["y2"])
            if natural_y1_label and natural_y2_label:
                _write_date_labels(ws, NATURAL_HVAC, natural_y1_label, natural_y2_label)

    # ── High Level Sheet ──
    if "High Level" in wb.sheetnames:
        ws = wb["High Level"]

        # PB delta (aggregate of all trades)
        hl_pre = data.get("high_level_pre", TradeMetrics())
        hl_post = data.get("high_level_post", TradeMetrics())
        _write_metrics(ws, HIGH_LEVEL_PB, hl_pre, hl_post)
        _write_date_labels(ws, HIGH_LEVEL_PB, pre_label, post_label)

        # Natural delta (aggregate) — sum natural y1/y2 across trades
        if data.get("natural"):
            nat_y1_total = TradeMetrics()
            nat_y2_total = TradeMetrics()
            for trade_nat in data["natural"].values():
                _add_trade_metrics(nat_y1_total, trade_nat["y1"])
                _add_trade_metrics(nat_y2_total, trade_nat["y2"])
            _write_metrics(ws, HIGH_LEVEL_NATURAL, nat_y1_total, nat_y2_total)
            if natural_y1_label and natural_y2_label:
                _write_date_labels(ws, HIGH_LEVEL_NATURAL, natural_y1_label, natural_y2_label)

    # ── Drive Time Sheet ──
    if "Drive Time" in wb.sheetnames:
        ws = wb["Drive Time"]
        dt_pre = data.get("drive_time_pre", DriveTimeMetrics())
        dt_post = data.get("drive_time_post", DriveTimeMetrics())
        _write_drive_time(ws, dt_pre, dt_post)
        _write_date_labels(ws, DRIVE_TIME_MAP, pre_label, post_label)

    wb.save(output_path)
    return str(output_path)


def _add_trade_metrics(target: TradeMetrics, source: TradeMetrics):
    """Accumulate source into target."""
    target.total_sales += source.total_sales
    target.total_tech_lead_sales += source.total_tech_lead_sales
    target.completed_revenue += source.completed_revenue
    target.completed_jobs += source.completed_jobs
    target.opportunities += source.opportunities
    target.sales_opportunities += source.sales_opportunities
    target.leads_set += source.leads_set
    target.converted_jobs += source.converted_jobs
