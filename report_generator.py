"""
Report Generator — Produces formatted Excel reports for 10+ Tag audits.

Tabs:
  1. "{Customer} 10+ Detail" — per-job breakdown with color-coded rows
  2. "Summary"               — aggregate counts and color legend
  3. "QA"                    — ServiceTitan QA results (optional)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper import JobRecord, QARecord


# ─── Colors ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Legend uses slightly different yellow
LEGEND_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CATEGORY_FILLS = {
    "Dispatcher placed 10+ tag that Probook Missed": RED_FILL,
    "Probook placed 10+ tag that CSR/Dispatch Missed": YELLOW_FILL,
    "10+ Priority Mismatch": ORANGE_FILL,
    "10+ Job Type Mismatch": BLUE_FILL,
    "10+ Tag Mismatch": GREEN_FILL,
}

# Category display order for summary
CATEGORY_ORDER = [
    "Dispatcher placed 10+ tag that Probook Missed",
    "Probook placed 10+ tag that CSR/Dispatch Missed",
    "10+ Priority Mismatch",
    "10+ Job Type Mismatch",
    "10+ Tag Mismatch",
    "Match",
]


# ─── Main Entry Point ────────────────────────────────────────────────────────

def generate_report(
    jobs: list[JobRecord],
    customer_name: str,
    output_dir: str = ".",
    qa_records: Optional[list[QARecord]] = None,
) -> str:
    """Generate the 10+ Tag Report Excel file."""
    relevant_jobs = [j for j in jobs if j.ai_has_10plus or j.disp_has_10plus]
    # Sort by category order
    cat_sort = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    relevant_jobs.sort(key=lambda j: cat_sort.get(j.category, 99))

    wb = Workbook()

    _build_report_tab(wb, relevant_jobs, customer_name)
    _build_summary_tab(wb, relevant_jobs, customer_name)

    if qa_records:
        _build_qa_tab(wb, qa_records)

    safe_name = customer_name.replace(" ", "_")
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{safe_name}_10Plus_Report_{date_str}.xlsx"
    filepath = Path(output_dir) / filename
    wb.save(str(filepath))

    return str(filepath)


# ─── Tab 1: Detail ───────────────────────────────────────────────────────────

REPORT_COLUMNS = [
    "Job ID",
    "Reasons",
    "ST Link",
    "Category",
    "VP Business Unit",
    "VP Job Type",
    "VP Priority",
    "Disp Job Type",
    "Probook Tags",
    "Disp Business Unit",
    "Disp Priority",
    "Disp Tags",
    "VP 10+ Tags",
    "Disp 10+ Tags",
]

# Column widths matching reference file
COLUMN_WIDTHS = [14, 98, 44, 60, 35, 30, 10, 35, 50, 35, 10, 87, 25, 25]


def _build_report_tab(wb: Workbook, jobs: list[JobRecord], customer_name: str) -> None:
    ws = wb.active
    ws.title = f"{customer_name} 10+ Detail"

    # Header row
    for col_idx, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        ai_10plus = ", ".join(t for t in job.ai_prediction.tags if "10+" in t) or "None"
        disp_10plus = ", ".join(t for t in job.dispatcher_verified.tags if "10+" in t) or "None"
        st_link = f"https://go.servicetitan.com/#/Job/Index/{job.job_id}"

        values = [
            int(job.job_id) if job.job_id.isdigit() else job.job_id,
            job.hvac_system_age_reason,
            st_link,
            job.category,
            job.business_unit,
            job.ai_prediction.job_type,
            job.ai_prediction.priority,
            job.dispatcher_verified.job_type,
            ", ".join(job.ai_prediction.tags),
            job.business_unit,
            job.dispatcher_verified.priority,
            ", ".join(job.dispatcher_verified.tags),
            ai_10plus,
            disp_10plus,
        ]

        row_fill = CATEGORY_FILLS.get(job.category)

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill:
                cell.fill = row_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if jobs:
        last_col = get_column_letter(len(REPORT_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(jobs) + 1}"

    # Set column widths
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─── Tab 2: Summary ──────────────────────────────────────────────────────────

def _build_summary_tab(wb: Workbook, jobs: list[JobRecord], customer_name: str) -> None:
    ws = wb.create_sheet("Summary")

    # Count per category
    counts = {}
    for cat in CATEGORY_ORDER:
        counts[cat] = sum(1 for j in jobs if j.category == cat)

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{customer_name} 10+ Mismatch Summary")
    # Row 2: blank
    # Row 3: Header
    for col, val in [(1, "Category"), (2, "Count"), (3, "Color")]:
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Rows 4-9: categories
    summary_fills = {
        "Dispatcher placed 10+ tag that Probook Missed": RED_FILL,
        "Probook placed 10+ tag that CSR/Dispatch Missed": YELLOW_FILL,
        "10+ Priority Mismatch": ORANGE_FILL,
        "10+ Job Type Mismatch": BLUE_FILL,
        "10+ Tag Mismatch": GREEN_FILL,
        "Match": None,
    }
    summary_labels = {
        "Dispatcher placed 10+ tag that Probook Missed": "MISSED 10+ (Probook under-classified)",
        "Probook placed 10+ tag that CSR/Dispatch Missed": "Probook placed 10+ tag that CSR/Dispatch Missed",
        "10+ Priority Mismatch": "10+ Priority Mismatch",
        "10+ Job Type Mismatch": "10+ Job Type Mismatch",
        "10+ Tag Mismatch": "10+ Tag Mismatch",
        "Match": "Match (Both sides agree)",
    }

    for i, cat in enumerate(CATEGORY_ORDER):
        row = 4 + i
        ws.cell(row=row, column=1, value=summary_labels[cat]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=counts[cat]).border = THIN_BORDER
        ws.cell(row=row, column=3).border = THIN_BORDER

    # Row 10: TOTAL
    total_row = 4 + len(CATEGORY_ORDER)
    ws.cell(row=total_row, column=1, value="TOTAL").border = THIN_BORDER
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row - 1})")
    total_cell.border = THIN_BORDER
    total_cell.font = Font(bold=True)

    # Blank row
    # Row 12: Color Legend header
    legend_start = total_row + 2
    ws.cell(row=legend_start, column=1, value="Color Legend:").font = Font(bold=True)

    legend_items = [
        (RED_FILL, "Red — MISSED 10+: Probook did not classify as 10+ but Dispatcher confirmed 10+"),
        (LEGEND_YELLOW, "Yellow — Priority Mismatch: Both agree on 10+ Job Type but Priority differs"),
        (ORANGE_FILL, "Orange — Over-aged: Probook said 10+ but Dispatcher had different non-10+ Job Type"),
        (BLUE_FILL, "Blue — Job Type Mismatch: Both 10+ but different specific 10+ Job Type"),
        (GREEN_FILL, "Green — Tag differences or tag-only 10+ references"),
    ]

    for i, (fill, text) in enumerate(legend_items):
        row = legend_start + 1 + i
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = fill

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10


# ─── Tab 3: QA ───────────────────────────────────────────────────────────────

QA_COLUMNS = ["Job ID", "Equipment Found", "Ages Found", "Screenshot Path", "Notes"]


def _build_qa_tab(wb: Workbook, qa_records: list[QARecord]) -> None:
    ws = wb.create_sheet("QA")

    for col_idx, header in enumerate(QA_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, qa in enumerate(qa_records, start=2):
        values = [
            qa.job_id,
            "Yes" if qa.equipment_found else "No",
            ", ".join(qa.ages_found) if qa.ages_found else "None found",
            qa.screenshot_path,
            qa.notes,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(QA_COLUMNS, start=1):
        width = min(max(len(header) + 4, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
