#!/usr/bin/env python3
"""
AI Val Recapture Report — Batch runner for Midwest brands.

Runs AI Validation for all Midwest brands, applies QA filtering,
and generates a single master spreadsheet.

Report criteria:
  - Missed 10+ Maintenances (excluding Service 10+ → Maintenance 10+)
  - Missed 10+ Services
  - Missed PL - Water Heater 8+

QA filtering (no cheap wins):
  - No First Time Customer / Unknown Age cases
  - No cases where dispatcher has a defensible position

Output columns:
  Date | Brand | CSR/DSR Verification Result | AI Validator Result |
  Reason for Change | Link to ST

Usage:
    python recapture_report.py --start-date 2026-02-01 --end-date 2026-02-28
"""

import argparse
import json
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper import ProBookScraper, JobRecord

# ─── Midwest Brands ─────────────────────────────────────────────────────────

MIDWEST_BRANDS = [
    "AB May",
    "Holt",
    "ServiceOne",
    "KB Complete",
    "Air Services",
    "Academy Morrison",
    "Haley",
    "Davison/Cregger",
    "Meridian",
    "Canfield",
    "Hero",
    "ASP",
    "Swick",
    "PRT",
    "West Allis",
    "Hockers",
]

# ─── Recapture Categories ────────────────────────────────────────────────────

CATEGORY_MISSED_MAINT_10PLUS = "Missed 10+ Maintenance"
CATEGORY_MISSED_SERVICE_10PLUS = "Missed 10+ Service"
CATEGORY_MISSED_PL_WH_8PLUS = "Missed PL - Water Heater 8+"

# ─── Excel Styles ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

MAINT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SERVICE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
WH_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CATEGORY_FILLS = {
    CATEGORY_MISSED_MAINT_10PLUS: MAINT_FILL,
    CATEGORY_MISSED_SERVICE_10PLUS: SERVICE_FILL,
    CATEGORY_MISSED_PL_WH_8PLUS: WH_FILL,
}

REPORT_COLUMNS = [
    "Date",
    "Brand",
    "CSR/DSR Verification Result",
    "AI Validator Result",
    "Reason for Change",
    "Link to ST",
    "Category",
    "Job ID",
]

ST_BASE_URL = "https://go.servicetitan.com/Job/Index"


# ─── QA Filtering ───────────────────────────────────────────────────────────

def is_first_time_customer(job: JobRecord) -> bool:
    """Check if the job is a First Time Customer / unknown age — cheap win."""
    # First call flag from either side
    if job.ai_prediction.is_first_call is True:
        return True
    if job.dispatcher_verified.is_first_call is True:
        return True
    return False


def is_unknown_age(job: JobRecord) -> bool:
    """Check if the job has unknown age — cheap win."""
    return job.unknown_age


def dispatcher_has_defensible_position(job: JobRecord) -> bool:
    """Check if the dispatcher's tagging is defensible.

    Example: AI says 10+ but dispatcher tagged based on a summary
    that said "5-10 years" — DSR put 5-9, which is reasonable.

    We check if the dispatcher tags contain age-related tags that
    are close to the 10+ threshold (e.g., 8-9, 5-10 range).
    """
    disp_tags = [t.lower() for t in job.dispatcher_verified.tags]
    ai_tags = [t.lower() for t in job.ai_prediction.tags]

    # If dispatcher has tags suggesting near-10+ age range, it's defensible
    # Common near-threshold tags: "5-10", "8-9", "5-9 years"
    near_threshold_patterns = [
        "5-10", "5-9", "8-9", "7-10", "6-10",
        "5 to 10", "5 to 9", "8 to 9",
    ]

    for tag in disp_tags:
        for pattern in near_threshold_patterns:
            if pattern in tag:
                # Check if AI is flagging 10+ — if dispatcher's range
                # overlaps or is close, it's defensible
                ai_has_10plus = any("10+" in t for t in ai_tags)
                if ai_has_10plus:
                    return True

    return False


def is_service_to_maintenance_reclassification(job: JobRecord) -> bool:
    """Check if the diff is just Service 10+ → Maintenance 10+ reclassification.

    This is NOT a real missed opportunity — just a job type change.
    Both sides have 10+, it's only the job type that changed.
    """
    ai_type = job.ai_prediction.job_type.lower()
    disp_type = job.dispatcher_verified.job_type.lower()

    ai_has_10 = any("10+" in t for t in job.ai_prediction.tags)
    disp_has_10 = any("10+" in t for t in job.dispatcher_verified.tags)

    # Service 10+ → Maintenance 10+ (both have 10+, just different job type)
    if "service" in disp_type and "maintenance" in ai_type and ai_has_10 and disp_has_10:
        return True
    return False


def passes_qa(job: JobRecord) -> bool:
    """Return True if the job passes QA (is NOT a cheap win)."""
    if is_first_time_customer(job):
        return False
    if is_unknown_age(job):
        return False
    if dispatcher_has_defensible_position(job):
        return False
    return True


# ─── Categorization ─────────────────────────────────────────────────────────

def categorize_job(job: JobRecord) -> Optional[str]:
    """Determine the recapture category for a job. Returns None if not applicable."""
    ai_type = job.ai_prediction.job_type.lower()
    disp_type = job.dispatcher_verified.job_type.lower()
    ai_tags = [t.lower() for t in job.ai_prediction.tags]
    disp_tags = [t.lower() for t in job.dispatcher_verified.tags]

    ai_has_10plus = any("10+" in t for t in ai_tags)
    disp_has_10plus = any("10+" in t for t in disp_tags)

    # Only interested in cases where AI flagged 10+ but dispatcher didn't
    # (AI Added 10+ = missed opportunity by dispatcher)
    if not ai_has_10plus or disp_has_10plus:
        return None

    # Check for PL Water Heater 8+ first (more specific)
    ai_has_8plus = any("8+" in t for t in ai_tags)
    is_plumbing = any(kw in ai_type for kw in ["plumb", "water heater", "wh"])
    is_plumbing_bu = any(kw in job.business_unit.lower() for kw in ["plumb", "water"])
    is_water_heater = any("water heater" in t or "wh" in t for t in ai_tags)

    if (is_plumbing or is_plumbing_bu or is_water_heater) and ai_has_8plus:
        return CATEGORY_MISSED_PL_WH_8PLUS

    # Missed 10+ Maintenance (exclude Service→Maintenance reclassification)
    if "maintenance" in ai_type and ai_has_10plus:
        if is_service_to_maintenance_reclassification(job):
            return None
        return CATEGORY_MISSED_MAINT_10PLUS

    # Missed 10+ Service
    if "service" in ai_type and ai_has_10plus:
        return CATEGORY_MISSED_SERVICE_10PLUS

    # Catch-all for other 10+ misses (install, etc.)
    if ai_has_10plus:
        # Default to service if job type is unclear
        if "maintenance" in ai_type or "maint" in ai_type:
            return CATEGORY_MISSED_MAINT_10PLUS
        return CATEGORY_MISSED_SERVICE_10PLUS

    return None


# ─── Reason Generation ──────────────────────────────────────────────────────

def generate_reason(job: JobRecord) -> str:
    """Generate a human-readable reason for the AI vs. dispatcher difference."""
    ai_tags = ", ".join(job.ai_prediction.tags) if job.ai_prediction.tags else "None"
    disp_tags = ", ".join(job.dispatcher_verified.tags) if job.dispatcher_verified.tags else "None"

    parts = []

    # Job type diff
    if job.ai_prediction.job_type != job.dispatcher_verified.job_type:
        parts.append(
            f"Job Type: DSR={job.dispatcher_verified.job_type} → AI={job.ai_prediction.job_type}"
        )

    # Tag diff
    ai_10plus_tags = [t for t in job.ai_prediction.tags if "10+" in t or "8+" in t]
    disp_10plus_tags = [t for t in job.dispatcher_verified.tags if "10+" in t or "8+" in t]

    if ai_10plus_tags and not disp_10plus_tags:
        parts.append(f"AI flagged: {', '.join(ai_10plus_tags)} — DSR did not tag")

    if not parts:
        parts.append(f"AI Tags: {ai_tags} | DSR Tags: {disp_tags}")

    return "; ".join(parts)


# ─── Extract Job Date ────────────────────────────────────────────────────────

def extract_job_date_from_scraper(scraper: ProBookScraper, job: JobRecord) -> str:
    """Try to extract the job date from the ProBook page for this job.

    Falls back to the dataset date range if individual dates aren't available.
    """
    # The job cards may have date info — try to get it from the page
    # If not available, use the dataset start date as a fallback
    try:
        date_text = scraper.page.evaluate(f"""
            () => {{
                const cards = document.querySelectorAll('div.border.rounded-md.p-4');
                for (const card of cards) {{
                    const text = card.textContent || '';
                    if (text.includes('Job ID: {job.job_id}')) {{
                        // Look for date patterns in the card
                        const dateMatch = text.match(/(\\d{{1,2}}\\/\\d{{1,2}}\\/\\d{{4}})|(\\d{{4}}-\\d{{2}}-\\d{{2}})/);
                        if (dateMatch) return dateMatch[0];
                    }}
                }}
                return '';
            }}
        """)
        if date_text:
            return date_text
    except Exception:
        pass

    return ""  # Will use date range as fallback


# ─── Master Report Generator ────────────────────────────────────────────────

def generate_master_report(
    all_records: list[dict],
    output_dir: str = ".",
    date_range: str = "",
) -> str:
    """Generate the master recapture spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Val Recapture Report"

    # Header row
    for col_idx, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, record in enumerate(all_records, start=2):
        values = [
            record["date"],
            record["brand"],
            record["csr_dsr_result"],
            record["ai_result"],
            record["reason"],
            record["st_link"],
            record["category"],
            record["job_id"],
        ]

        row_fill = CATEGORY_FILLS.get(record["category"])

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill:
                cell.fill = row_fill

            # Make ST link a hyperlink
            if col_idx == 6 and value:
                cell.font = Font(color="0563C1", underline="single")

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    if all_records:
        last_col = get_column_letter(len(REPORT_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(all_records) + 1}"

    # Column widths
    widths = [14, 20, 30, 30, 50, 45, 25, 12]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary sheet
    _build_summary_sheet(wb, all_records)

    # Save
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"Midwest_AI_Val_Recapture_{timestamp}.xlsx"
    filepath = Path(output_dir) / filename
    wb.save(str(filepath))
    return str(filepath)


def _build_summary_sheet(wb: Workbook, records: list[dict]) -> None:
    """Build a summary breakdown by brand and category."""
    ws = wb.create_sheet("Summary")

    # Title
    ws.cell(row=1, column=1, value="AI Val Recapture Summary").font = Font(bold=True, size=14)

    # Headers
    headers = ["Brand", CATEGORY_MISSED_MAINT_10PLUS, CATEGORY_MISSED_SERVICE_10PLUS,
               CATEGORY_MISSED_PL_WH_8PLUS, "Total"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Count by brand
    brand_counts = {}
    for r in records:
        brand = r["brand"]
        cat = r["category"]
        if brand not in brand_counts:
            brand_counts[brand] = {
                CATEGORY_MISSED_MAINT_10PLUS: 0,
                CATEGORY_MISSED_SERVICE_10PLUS: 0,
                CATEGORY_MISSED_PL_WH_8PLUS: 0,
            }
        brand_counts[brand][cat] += 1

    row = 4
    for brand in MIDWEST_BRANDS:
        counts = brand_counts.get(brand, {
            CATEGORY_MISSED_MAINT_10PLUS: 0,
            CATEGORY_MISSED_SERVICE_10PLUS: 0,
            CATEGORY_MISSED_PL_WH_8PLUS: 0,
        })
        total = sum(counts.values())
        values = [brand, counts[CATEGORY_MISSED_MAINT_10PLUS],
                  counts[CATEGORY_MISSED_SERVICE_10PLUS],
                  counts[CATEGORY_MISSED_PL_WH_8PLUS], total]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.border = THIN_BORDER
        row += 1

    # Grand total
    grand_total = len(records)
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=5, value=grand_total).font = Font(bold=True)
    for col_idx in range(1, 6):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 28


# ─── Batch Runner ────────────────────────────────────────────────────────────

def process_brand(
    scraper: ProBookScraper,
    brand: str,
    start_date: str,
    end_date: str,
) -> list[dict]:
    """Run the full pipeline for one brand and return filtered recapture records."""
    records = []

    try:
        # Select customer
        scraper.customer_name = brand
        scraper._select_customer()

        # Navigate to Audit
        scraper.navigate_to_audit()

        # Build Dataset
        scraper.start_date = start_date
        scraper.end_date = end_date
        scraper.build_dataset()

        # Run Validation — MUST wait for 100%
        scraper.run_validation()

        # Navigate to Diffs, scroll, scrape
        scraper.navigate_to_diffs()

        # Extra wait for diffs to fully load (user specified 10-20 min)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Waiting for diffs to fully load...")
        time.sleep(15 * 60)  # 15 minutes

        scraper.scroll_and_load_all_jobs()
        jobs = scraper.scrape_all_jobs()

        # Process each job
        for job in jobs:
            # Categorize
            category = categorize_job(job)
            if category is None:
                continue

            # QA filter — skip cheap wins
            if not passes_qa(job):
                continue

            # Try to get job date
            job_date = extract_job_date_from_scraper(scraper, job)
            if not job_date:
                job_date = f"{start_date} - {end_date}"

            # Build CSR/DSR result string
            disp_result_parts = []
            if job.dispatcher_verified.job_type:
                disp_result_parts.append(job.dispatcher_verified.job_type)
            if job.dispatcher_verified.tags:
                disp_result_parts.append(", ".join(job.dispatcher_verified.tags))
            csr_dsr_result = " | ".join(disp_result_parts) if disp_result_parts else "No tags"

            # Build AI result string
            ai_result_parts = []
            if job.ai_prediction.job_type:
                ai_result_parts.append(job.ai_prediction.job_type)
            if job.ai_prediction.tags:
                ai_result_parts.append(", ".join(job.ai_prediction.tags))
            ai_result = " | ".join(ai_result_parts) if ai_result_parts else "No tags"

            records.append({
                "date": job_date,
                "brand": brand,
                "csr_dsr_result": csr_dsr_result,
                "ai_result": ai_result,
                "reason": generate_reason(job),
                "st_link": f"{ST_BASE_URL}/{job.job_id}",
                "category": category,
                "job_id": job.job_id,
            })

        print(f"[{datetime.now().strftime('%H:%M:%S')}] {brand}: {len(records)} recapture records after QA filtering")

    except Exception as e:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR processing {brand}: {e}")

    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="AI Val Recapture Report — Midwest Batch")
    parser.add_argument("--start-date", required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", required=True, help="End date (YYYY-MM-DD)")
    parser.add_argument("--output-dir", default=".", help="Output directory")
    parser.add_argument(
        "--brands",
        nargs="*",
        default=None,
        help="Specific brands to run (default: all Midwest brands)",
    )
    parser.add_argument(
        "--no-interactive",
        action="store_true",
        default=False,
        help="Run without input() pauses",
    )
    args = parser.parse_args()

    brands = args.brands if args.brands else MIDWEST_BRANDS
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("  AI VAL RECAPTURE REPORT — MIDWEST")
    print(f"  Date Range: {args.start_date} to {args.end_date}")
    print(f"  Brands: {len(brands)}")
    print("=" * 60)

    all_records = []

    # Launch browser once
    scraper = ProBookScraper(
        customer_name=brands[0],
        start_date=args.start_date,
        end_date=args.end_date,
        interactive=not args.no_interactive,
    )

    try:
        scraper.launch()
        scraper.login_and_select_customer()

        # Process first brand (already selected)
        print(f"\n{'─' * 40}")
        print(f"  Processing 1/{len(brands)}: {brands[0]}")
        print(f"{'─' * 40}")

        # Navigate to Audit for first brand
        scraper.navigate_to_audit()
        scraper.build_dataset()
        scraper.run_validation()
        scraper.navigate_to_diffs()

        print(f"[{datetime.now().strftime('%H:%M:%S')}] Waiting for diffs to fully load (15 min)...")
        time.sleep(15 * 60)

        scraper.scroll_and_load_all_jobs()
        jobs = scraper.scrape_all_jobs()

        # Process jobs for first brand
        brand_records = _process_jobs(jobs, brands[0], scraper, args.start_date, args.end_date)
        all_records.extend(brand_records)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {brands[0]}: {len(brand_records)} recapture records")

        # Save JSON backup for first brand
        scraper.save_json_backup(output_dir=str(output_dir))

        # Process remaining brands
        for idx, brand in enumerate(brands[1:], start=2):
            print(f"\n{'─' * 40}")
            print(f"  Processing {idx}/{len(brands)}: {brand}")
            print(f"{'─' * 40}")

            try:
                # Navigate back to dashboard/home to switch customer
                scraper.page.goto(scraper.PROBOOK_URL, wait_until="networkidle")
                scraper.page.wait_for_timeout(3000)

                # Update scraper state
                scraper.customer_name = brand
                scraper.start_date = args.start_date
                scraper.end_date = args.end_date
                scraper._dataset_name = ""
                scraper.jobs = []

                # Select new customer
                scraper._select_customer()
                scraper.page.wait_for_timeout(2000)

                # Run the pipeline
                scraper.navigate_to_audit()
                scraper.build_dataset()
                scraper.run_validation()
                scraper.navigate_to_diffs()

                print(f"[{datetime.now().strftime('%H:%M:%S')}] Waiting for diffs to fully load (15 min)...")
                time.sleep(15 * 60)

                scraper.scroll_and_load_all_jobs()
                jobs = scraper.scrape_all_jobs()

                # Process and filter
                brand_records = _process_jobs(jobs, brand, scraper, args.start_date, args.end_date)
                all_records.extend(brand_records)
                print(f"[{datetime.now().strftime('%H:%M:%S')}] {brand}: {len(brand_records)} recapture records")

                # Save JSON backup
                scraper.save_json_backup(output_dir=str(output_dir))

            except Exception as e:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] ERROR on {brand}: {e}")
                print(f"  Saving partial data and continuing to next brand...")

                # Try to save partial data
                if scraper.jobs:
                    try:
                        scraper.save_json_backup(output_dir=str(output_dir))
                    except Exception:
                        pass

    except KeyboardInterrupt:
        print("\nInterrupted by user. Generating report with data collected so far...")

    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        print("Generating report with data collected so far...")

    finally:
        scraper.close()

    # Generate master report
    if all_records:
        report_path = generate_master_report(
            all_records,
            output_dir=str(output_dir),
            date_range=f"{args.start_date} to {args.end_date}",
        )
        print(f"\n{'=' * 60}")
        print(f"  MASTER REPORT SAVED: {report_path}")
        print(f"  Total recapture records: {len(all_records)}")
        print(f"{'=' * 60}")
    else:
        print("\nNo recapture records found across all brands.")

    # Save all records as JSON backup
    json_path = output_dir / f"recapture_all_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    json_path.write_text(json.dumps(all_records, indent=2, default=str))
    print(f"JSON backup: {json_path}")


def _process_jobs(
    jobs: list[JobRecord],
    brand: str,
    scraper: ProBookScraper,
    start_date: str,
    end_date: str,
) -> list[dict]:
    """Filter and transform jobs into recapture records for one brand."""
    records = []

    for job in jobs:
        # Categorize
        category = categorize_job(job)
        if category is None:
            continue

        # QA filter
        if not passes_qa(job):
            continue

        # Try to get job date
        job_date = extract_job_date_from_scraper(scraper, job)
        if not job_date:
            job_date = f"{start_date} - {end_date}"

        # CSR/DSR result
        disp_parts = []
        if job.dispatcher_verified.job_type:
            disp_parts.append(job.dispatcher_verified.job_type)
        if job.dispatcher_verified.tags:
            disp_parts.append(", ".join(job.dispatcher_verified.tags))
        csr_dsr_result = " | ".join(disp_parts) if disp_parts else "No tags"

        # AI result
        ai_parts = []
        if job.ai_prediction.job_type:
            ai_parts.append(job.ai_prediction.job_type)
        if job.ai_prediction.tags:
            ai_parts.append(", ".join(job.ai_prediction.tags))
        ai_result = " | ".join(ai_parts) if ai_parts else "No tags"

        records.append({
            "date": job_date,
            "brand": brand,
            "csr_dsr_result": csr_dsr_result,
            "ai_result": ai_result,
            "reason": generate_reason(job),
            "st_link": f"{ST_BASE_URL}/{job.job_id}",
            "category": category,
            "job_id": job.job_id,
        })

    return records


if __name__ == "__main__":
    main()
